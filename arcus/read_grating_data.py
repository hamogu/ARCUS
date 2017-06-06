import numpy as np
import astropy.units as u
from scipy.interpolate import RectBivariateSpline
from openpyxl import load_workbook
from marxs.base import SimulationSequenceElement


class DataFileFormatException(Exception):
    pass


class InterpolateRalfTable(object):
    '''Order Selector for MARXS using a specific kind of data table.

    The data table was given to me by Ralf. It contains simulated grating
    efficiencies in an Excel table.
    A short summary of this format is given here, to help reading the code.
    The table contains data in 3-dimenasional (wave, n_theta, order) space,
    flattened into a 2d table.

    - Row 1 + 2: Column labels. Not used here.
    - Column A: wavelength in nm.
    - Column B: blaze angle in deg.
    - Rest: data

    For each wavelength there are multiple blaze angles listed, so Column A
    contains
    many dublicates and looks like this: [1,1,1,1,1,1,2,2,2,2,2,2,3,3,3, ...].
    Column B repeats like this: [1,2,3,4,5,6,1,2,3,4,5,6,1,2,3, ...].

    Because the wave, theta grid is regular, this class can use the
    `scipy.interpolate.RectBivariateSpline` for interpolation in each 2-d slice
    (``order`` is an integer and not interpolated).

    Parameters
    ----------
    filename : string
        path and name of data file
    k : int
        Degree of spline. See `scipy.interpolate.RectBivariateSpline`.
    '''

    def __init__(self, filename, k=3):
        ralf = load_workbook(filename)
        sheet = ralf.get_sheet_by_name('Sheet1')
        wave = np.array([d[0].value for d in sheet['A3:A{0}'.format(sheet.max_row)]])
        theta = np.array([d[0].value for d in sheet['B3:B{0}'.format(sheet.max_row)]])
        self.orders = -np.array([d.value for d in list(sheet.get_squared_range(3, 2, sheet.max_column, 2))[0]])
        datagenerator = sheet.get_squared_range(3, 3, sheet.max_column, sheet.max_row)
        data = np.empty((sheet.max_column - 2, sheet.max_row - 2))
        for i, r in enumerate(datagenerator):
            for j, d in enumerate(r):
                data[j, i] = d.value

        n_wave = len(set(wave))
        n_theta = len(set(theta))
        if data.shape != (len(self.orders), n_wave * n_theta):
            raise DataFileFormatException('Check file format.')

        # This really is 3d data in (wave, n_theta, order) space.
        # Need to fold back into 3d space.

        wave = wave[::n_theta]
        theta = np.deg2rad(theta[:n_theta])
        self.data = data.reshape((len(self.orders), n_wave, n_theta))

        # Order is int, we will never interpolate about order, thus, we'll just have
        # len(order) 2d interpolations
        self.interpolators = [RectBivariateSpline(wave, theta, self.data[i, :,:], kx=k, ky=k) for i in range(len(self.orders))]

    def probabilities(self, energies, pol, blaze):
        '''Obtain the probabilties for photons to go into a particular order.

        This has the same parameters as the ``__call__`` method, but it returns
        the raw probabilities, while ``__call__`` will draw from these
        probabilities and assign an order and a total survival probability to
        each photon.

        Parameters
        ----------
        energies : np.array
            Energy for each photons
        pol : np.array
            Polarization for each photon (not used in this class)
        blaze : np.array
            Blaze angle for each photon

        Returns
        -------
        orders : np.array
            Array of orders
        interpprobs : np.array
            This array contains a probability array for each photon to reach a
            particular order
        '''
        # convert energy in keV to wavelength in nm
        # (nm is the unit of the input table)
        wave = (energies * u.keV).to(u.nm, equivalencies=u.spectral()).value
        interpprobs = np.empty((len(self.orders), len(energies)))
        for i, interp in enumerate(self.interpolators):
            interpprobs[i, :] = interp.ev(wave, blaze)
        return self.orders, interpprobs

    def __call__(self, energies, pol, blaze):
        orders, interpprobs = self.probabilities(energies, pol, blaze)
        totalprob = np.sum(interpprobs, axis=0)
        # Cumulative probability for orders, normalized to 1.
        cumprob = np.cumsum(interpprobs, axis=0) / totalprob
        ind_orders = np.argmax(cumprob > np.random.rand(len(energies)), axis=0)

        return orders[ind_orders], totalprob


class RalfQualityFactor(SimulationSequenceElement):
    '''Scale probabilites of theoretical curves to measured values.

    All gratings look better in theory than in practice. This grating quality factor
    scales the calculated diffraction probabilities to the observed performance.
    '''

    def __init__(self, **kwargs):
        self.sigma = kwargs.pop('sigma')
        self.d = kwargs.pop('d')
        super(RalfQualityFactor, self).__init__(**kwargs)

    def process_photons(self, photons):
        ind = np.isfinite(photons['order'])
        photons['probability'][ind] *= np.exp(- (2 * np.pi * self.sigma / self.d)**2)**(photons['order'][ind]**2)
        return photons
